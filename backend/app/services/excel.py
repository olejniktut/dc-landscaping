from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List
from datetime import date
from decimal import Decimal

from app.models.time_record import TimeRecord


def create_report_excel(
    records: List[TimeRecord],
    start_date: date,
    end_date: date,
    property_name: str = "All Properties"
) -> BytesIO:
    """Create an Excel report from time records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_font = Font(bold=True)
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"DC Landscaping - Time Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Info
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Period: {start_date} to {end_date} | Property: {property_name}"
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Date", "Property", "Type", "Workers", "Hours", "Cost"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    total_hours = Decimal("0")
    total_cost = Decimal("0")
    
    row = 5
    for record in records:
        # Determine type
        cleanup_type = ""
        if record.property.is_spring_cleanup:
            cleanup_type = "Spring"
        elif record.property.is_fall_cleanup:
            cleanup_type = "Fall"
        
        # Calculate hours
        hours = Decimal(str(record.total_minutes / 60)) if record.total_minutes else Decimal("0")
        cost = record.total_cost or Decimal("0")
        
        total_hours += hours
        total_cost += cost
        
        # Worker names
        worker_names = ", ".join([w.name for w in record.workers])
        
        data = [
            record.work_date.strftime("%Y-%m-%d"),
            record.property.name,
            cleanup_type,
            worker_names,
            float(hours),
            float(cost)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 5:  # Hours
                cell.number_format = "0.00"
            elif col == 6:  # Cost
                cell.number_format = "0.00"
        
        row += 1
    
    # Totals row
    ws.cell(row=row, column=1, value="").border = border
    ws.cell(row=row, column=2, value="").border = border
    ws.cell(row=row, column=3, value="").border = border
    
    total_label = ws.cell(row=row, column=4, value="TOTAL:")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.border = border
    total_label.alignment = Alignment(horizontal="right")
    
    hours_cell = ws.cell(row=row, column=5, value=float(total_hours))
    hours_cell.font = total_font
    hours_cell.fill = total_fill
    hours_cell.border = border
    hours_cell.number_format = "0.00"
    
    cost_cell = ws.cell(row=row, column=6, value=float(total_cost))
    cost_cell.font = total_font
    cost_cell.fill = total_fill
    cost_cell.border = border
    cost_cell.number_format = "0.00"
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
